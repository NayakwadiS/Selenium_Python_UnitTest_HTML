import openpyxl
import os
import inspect
import re

def ReadLine(sFileName,sTestID):
    path =os.path.dirname(os.path.dirname(os.path.dirname(__file__)))+"\TestResources\TestData" +"\\"+sFileName
    wb = openpyxl.load_workbook(path, True)
    wsheet = wb.worksheets.__getitem__(0)

    #print("Sheet Title ", self.wsheet.title)
    #print("Totle col used ",self.wsheet.max_column)
    #print("Totle row used ",self.wsheet.max_row)
    #print("cell value 1,1 is",self.wsheet.cell(2,5).value)

    #for row in self.wsheet.rows:                                       #this for directly itterate over complet sheet and returen RAWCELL VALUE
     #   for cell in row:
      #      print(cell.value)

    #define dictioanary to store key-value pair of col and row data
    currentrow={}
    rownum =0
    for row in range(wsheet.max_row):                              #ittarate over all rows
        # for col in range(wsheet.max_column):                       #itterate over all col
            scellValue=wsheet.cell(row + 1, 1).value              #get value of TestId col.
            if len(re.findall(scellValue, sTestID)) > 0:
            # if(scellValue.find(sTestID,0,len(scellValue))):             #match value of TestID col to given test case ID
                rownum =wsheet.cell(row, 1).row
                # print("row",rownum)                                     #get row number
                break

    for col in range(wsheet.max_column):
        #print((wsheet.cell(1, col ).value))
        #print(col)
        #print((wsheet.cell(rownum + 1, col + 1).value))
        currentrow[wsheet.cell(1, col + 1).value] = wsheet.cell(rownum + 1,col + 1).value  # put col as key and row value ar 'value' in dict
    #print(currentrow)
    return currentrow

def getLastRow(wb):
    return wb.wsheet.max_row


#E1= ExcelReader()
#print("max row ",ExcelReader.getLastRow())
#currentrow=E1.ReadLine("SBI_blue_Chip")
#print(currentrow.keys())
#print(currentrow.values())
#print(currentrow.get('TestID'))